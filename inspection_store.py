"""
data_store.py — master lookup, Excel read/write, stats
"""
import json
import threading
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = Path(__file__).parent
MASTER_FILE = BASE_DIR / "Asset_Inventory.xlsx"
OUTPUT_FILE = BASE_DIR / "Inspection_Data.xlsx"
INSPECTORS_FILE = BASE_DIR / "inspectors.json"  # real names — gitignored, not committed


def _load_inspectors() -> list[str]:
    """Load inspector names from inspectors.json (local, gitignored — see README).

    Falls back to a small placeholder list so the app still runs out of
    the box without real names committed to source control.
    """
    if INSPECTORS_FILE.exists():
        try:
            with open(INSPECTORS_FILE, encoding="utf-8") as f:
                names = json.load(f)
            if isinstance(names, list) and names:
                return names
        except (json.JSONDecodeError, OSError):
            pass
    return ["Staff Member 1", "Staff Member 2", "Staff Member 3"]


INSPECTORS = _load_inspectors()

HEADERS = [
    "RECORD_ID", "DATE", "EQUIPMENT_SERIAL", "REGION", "SUB", "LOCATION",
    "FEEDER_UNIT", "INSPECTOR", "PRIMARY_SEAL_ID", "BACKUP_SEAL_ID",
    "COUNTER_A", "COUNTER_B", "STATUS", "TRANSMISSION",
    "CONFIG_DATE_CORRECT", "BASELINE_CHECK", "MULTI_ORIENTATION_SAMPLE_TEST",
    "RESULT_PATTERN_CONFIRMED", "BACKUP_MEDIA_SEALED", "MAIN_COMPARTMENT_SEALED",
    "ISSUE_DETAILS",
]

MASTER     = {}
_lock      = threading.Lock()

# Region targets — from Asset_Inventory.xlsx
# Re-run generate_region_targets.py if master file changes
REGION_TARGETS = {
    1: 52, 2: 68, 3: 83, 4: 65, 5: 64,
    6: 52, 7: 46, 8: 62, 9: 54, 10: 91,
    11: 72, 12: 62, 13: 88, 14: 50, 15: 54,
    16: 60, 17: 59, 18: 57, 19: 54, 20: 54,
    21: 46, 22: 49, 23: 38, 24: 48, 25: 49,
    'UNASSIGNED': 513,
}

# ── styles ─────────────────────────────────────────────────────────────────────
def _border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(status):
    c = {"COMPLETED": "E2EFDA", "FAILED": "FCE4D6", "UNDER_MAINTENANCE": "FFF2CC"}.get(status, "FFFFFF")
    return PatternFill("solid", fgColor=c)

def _cv(status):
    return "PASS" if status == "COMPLETED" else ("FAIL" if status == "FAILED" else "N/A")

def _tx(status):
    return "Sent" if status == "COMPLETED" else ("Failed" if status == "FAILED" else "Not Applicable")

# ── master lookup ──────────────────────────────────────────────────────────────
def _load_master():
    if not MASTER_FILE.exists():
        print(f"  WARNING: {MASTER_FILE.name} not found.")
        return {}
    print(f"  Reading {MASTER_FILE.name}...")
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True, keep_vba=False)
    ws = wb.active
    hdrs = {}
    data = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            hdrs = {str(c).strip().upper(): j for j, c in enumerate(row) if c}
            continue
        def gi(key):
            idx = hdrs.get(key.upper())
            return row[idx] if idx is not None and idx < len(row) else None
        sn = gi("ASSET_SN")
        if not sn or not str(sn).startswith("DS"):
            continue
        def sv(v):
            if v is None: return ""
            try: return int(v)
            except: return str(v).strip()
        data[str(sn).strip()] = {
            "region":    sv(gi("REGION")),
            "sub":     sv(gi("SUB")),
            "loc":     sv(gi("STG_LOC_AREA")),
            "sim":     sv(gi("SIM_NO")),
            "imei":    sv(gi("MODEM_IMEI")),
            "purpose": sv(gi("ASSET_PURPOSE")),
        }
    wb.close()
    print(f"  Loaded {len(data)} devices.")
    return data

# ── init excel ─────────────────────────────────────────────────────────────────
def _init_excel():
    if OUTPUT_FILE.exists():
        print(f"  Found existing {OUTPUT_FILE.name}")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection_Data"
    hf = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 2)
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"  Created {OUTPUT_FILE.name}")

# ── save record ────────────────────────────────────────────────────────────────
def save_record(data):
    s = data.get("status", "")
    cv, tx = _cv(s), _tx(s)
    with _lock:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["Inspection_Data"]
        rid = ws.max_row  # header is row 1
        row_vals = [
            rid,
            data.get("date", str(date.today())),
            data.get("serial", ""),
            data.get("region", ""),
            data.get("sub", ""),
            data.get("location", ""),
            data.get("feeder", ""),
            data.get("qa", ""),
            data.get("main_seal", ""),
            data.get("backup_seal", ""),
            data.get("counter_a", "") or "",
            data.get("counter_b", "") or "",
            s, tx,
            cv, cv, cv, cv, cv, cv,
            data.get("notes", ""),
        ]
        fill = _fill(s)
        brd  = _border()
        nr   = ws.max_row + 1
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=nr, column=col, value=val)
            c.fill = fill
            c.border = brd
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal="left" if col == len(HEADERS) else "center",
                vertical="center"
            )
        ws.row_dimensions[nr].height = 18
        try:
            wb.save(OUTPUT_FILE)
        except PermissionError:
            raise PermissionError("Inspection_Data.xlsx is currently open in Excel. Please close it and try again.")
    return rid

# ── find record ────────────────────────────────────────────────────────────────
def find_record(serial):
    if not OUTPUT_FILE.exists():
        return None
    serial = serial.strip().upper()
    wb   = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    ws   = wb["Inspection_Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    match = None
    for row in rows:
        if not row or len(row) < 3:
            continue
        cell_sn = row[2]
        if cell_sn and str(cell_sn).strip().upper() == serial:
            match = row

    if not match:
        return None

    def safe(idx):
        try: return match[idx] if match[idx] is not None else ""
        except: return ""

    return {
        "row_serial":      str(safe(2)).strip(),
        "record_id":       safe(0),
        "date":            str(safe(1)),
        "region":            safe(3),
        "sub":             safe(4),
        "location":        safe(5),
        "feeder":          str(safe(6)),
        "qa":              str(safe(7)),
        "main_seal":       str(safe(8)),
        "backup_seal":     str(safe(9)),
        "counter_a": str(safe(10)),
        "counter_b":    str(safe(11)),
        "status":          str(safe(12)),
        "notes":           str(safe(20)),
    }

# ── update record — adds a NEW ROW (keeps history) ────────────────────────────
def update_record(data):
    """
    Instead of overwriting the existing row, we ADD a new row so the full
    history is preserved. The original MAINTENANCE/FAILED row stays intact.
    """
    serial = str(data.get("serial", "")).strip().upper()
    s      = data.get("status", "")
    cv, tx = _cv(s), _tx(s)

    # First find the original record to copy region/sub/location from it
    original = find_record(serial)

    with _lock:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["Inspection_Data"]
        rid = ws.max_row  # next record ID

        row_vals = [
            rid,
            str(date.today()),
            serial,
            data.get("region",     original.get("region",    "") if original else ""),
            data.get("sub",      original.get("sub",     "") if original else ""),
            data.get("location", original.get("location","") if original else ""),
            data.get("feeder",   ""),
            data.get("qa",       ""),
            data.get("main_seal",    ""),
            data.get("backup_seal",  ""),
            data.get("counter_a", "") or "",
            data.get("counter_b",    "") or "",
            s, tx,
            cv, cv, cv, cv, cv, cv,
            data.get("notes", ""),
        ]

        fill = _fill(s)
        brd  = _border()
        nr   = ws.max_row + 1
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=nr, column=col, value=val)
            c.fill      = fill
            c.border    = brd
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal="left" if col == len(HEADERS) else "center",
                vertical="center"
            )
        ws.row_dimensions[nr].height = 18

        try:
            wb.save(OUTPUT_FILE)
        except PermissionError:
            raise PermissionError("Inspection_Data.xlsx is currently open in Excel. Please close it and try again.")

    return rid

# ── stats ──────────────────────────────────────────────────────────────────────
def get_stats():
    if not OUTPUT_FILE.exists():
        return {"total": 0, "completed": 0, "failed": 0, "maintenance": 0, "pass_rate": 0, "regions": {}}
    wb   = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    ws   = wb["Inspection_Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # ── Deduplicate by serial — keep the LATEST row per device ──────────────
    # Rows are ordered by RECORD_ID (ascending), so last occurrence = latest
    latest = {}  # serial -> row
    for row in rows:
        if not row or not row[0]: continue
        serial = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ""
        if serial:
            latest[serial] = row  # overwrite keeps last (latest) record

    total = completed = failed = maintenance = 0
    regions = {}

    for serial, row in latest.items():
        total += 1
        status   = str(row[12]).strip() if len(row) > 12 and row[12] else ""
        region_raw = row[3] if len(row) > 3 and row[3] else "Unknown"
        try:    region = str(int(region_raw))
        except: region = str(region_raw).strip()

        if status == "COMPLETED":           completed   += 1
        elif status == "FAILED":            failed      += 1
        elif status == "UNDER_MAINTENANCE": maintenance += 1

        if region not in regions:
            regions[region] = {"completed": 0, "failed": 0, "maintenance": 0,
                           "total": 0, "target": 0}
        regions[region]["total"] += 1
        if status == "COMPLETED":           regions[region]["completed"]   += 1
        elif status == "FAILED":            regions[region]["failed"]      += 1
        elif status == "UNDER_MAINTENANCE": regions[region]["maintenance"] += 1

    # ── Inject region targets ──────────────────────────────────────────────────
    for region_key, region_data in regions.items():
        try:    lookup = int(region_key)
        except: lookup = region_key
        region_data["target"] = REGION_TARGETS.get(lookup,
                              REGION_TARGETS.get(str(lookup), 0))

    # ── Overall pass rate against total target ───────────────────────────────
    total_target = sum(REGION_TARGETS.values()) if REGION_TARGETS else total
    pass_rate = round(completed / total_target * 100, 1) if total_target else 0

    return {
        "total": total, "completed": completed, "failed": failed,
        "maintenance": maintenance,
        "pass_rate": pass_rate,
        "has_targets": bool(REGION_TARGETS),
        "regions": regions,
    }



# ── search records ─────────────────────────────────────────────────────────────
def search_records(region="", sub=""):
    if not OUTPUT_FILE.exists():
        return []
    wb   = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    ws   = wb["Inspection_Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    results = []
    for row in rows:
        if not row or not row[0]: continue
        row_ward = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        row_sub  = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        # Normalise region comparison
        match_ward = not region or str(int(float(region))) == str(int(float(row_ward))) if region and row_ward else not region
        match_sub  = not sub  or row_sub == str(int(float(sub))) if sub else not sub

        if match_ward and match_sub:
            def safe(idx):
                try: return row[idx] if row[idx] is not None else ""
                except: return ""
            results.append({
                "row_serial":      str(safe(2)).strip(),
                "record_id":       safe(0),
                "date":            str(safe(1)),
                "region":            safe(3),
                "sub":             safe(4),
                "location":        safe(5),
                "feeder":          str(safe(6)),
                "qa":              str(safe(7)),
                "main_seal":       str(safe(8)),
                "backup_seal":     str(safe(9)),
                "counter_a": str(safe(10)),
                "counter_b":    str(safe(11)),
                "status":          str(safe(12)),
                "notes":           str(safe(20)),
            })
    return results

# ── daily stats ────────────────────────────────────────────────────────────────
def get_daily_stats():
    if not OUTPUT_FILE.exists():
        return []
    wb   = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    ws   = wb["Inspection_Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    from collections import defaultdict
    daily = defaultdict(int)
    for row in rows:
        if not row or not row[0]: continue
        status = str(row[12]).strip() if len(row) > 12 and row[12] else ""
        raw    = row[1] if len(row) > 1 else None
        if not raw: continue
        if hasattr(raw, 'strftime'):
            d = raw.strftime("%Y-%m-%d")
        else:
            d = str(raw).strip()
        if status == "COMPLETED" and d:
            daily[d] += 1
    return sorted([{"date": k, "completed": v} for k, v in daily.items()],
                  key=lambda x: x["date"])

# ── init ───────────────────────────────────────────────────────────────────────
def init():
    global MASTER
    MASTER = _load_master()
    _init_excel()
