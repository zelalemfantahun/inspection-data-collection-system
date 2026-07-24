"""
Toronto Elections — LAT Data Collection System
Run:  python app.py
"""
import os, secrets, threading, webbrowser
from functools import wraps
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import data_store

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)  # new random session key every process start

# The original code referenced APP_PASSWORD in a print() statement without
# ever defining it — a real NameError-on-startup bug (see SECURITY.md).
# Fixed: read from an environment variable if set, otherwise generate a
# random one-time password each run (same pattern used for Tab-Tech).
APP_PASSWORD = os.environ.get("LAT_APP_PASSWORD") or secrets.token_urlsafe(6)


def login_required(view):
    """The original login.html posted to /login, but no /login route or
    session check existed anywhere — every route was reachable without
    authentication despite a login page existing (see SECURITY.md)."""
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("authed"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form.get("password") == APP_PASSWORD:
            session["authed"] = True
            return redirect(url_for("index"))
        return render_template("login.html", error=True)
    return render_template("login.html", error=False)


@app.route("/logout")
def logout():
    session.pop("authed", None)
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    return render_template("form.html", staff=data_store.STAFF)

@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html", stats=data_store.get_stats())

@app.route("/update")
@login_required
def update_page():
    return render_template("update.html", staff=data_store.STAFF)

@app.route("/api/lookup/<serial>")
@login_required
def lookup(serial):
    d = data_store.MASTER.get(serial.strip().upper())
    if d:
        return jsonify({"found": True, **d})
    return jsonify({"found": False})

@app.route("/api/save", methods=["POST"])
@login_required
def save():
    try:
        rid = data_store.save_record(request.get_json())
        return jsonify({"ok": True, "id": rid})
    except PermissionError:
        return jsonify({"ok": False, "locked": True,
                        "error": "LAT_Data.xlsx is open in Excel. Please close it and try again."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/find/<serial>")
@login_required
def find(serial):
    try:
        rec = data_store.find_record(serial.strip().upper())
        if rec:
            return jsonify({"found": True, **rec})
        return jsonify({"found": False})
    except Exception as e:
        return jsonify({"found": False, "error": str(e)})

@app.route("/api/update", methods=["POST"])
@login_required
def update():
    try:
        data_store.update_record(request.get_json())
        return jsonify({"ok": True})
    except PermissionError:
        return jsonify({"ok": False, "locked": True,
                        "error": "LAT_Data.xlsx is open in Excel. Please close it and try again."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/debug-dates")
@login_required
def debug_dates():
    """Shows raw date values from Excel for troubleshooting."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(data_store.OUTPUT_FILE, read_only=True, data_only=True)
        ws = wb["LAT_Data"]
        rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
        wb.close()
        result = []
        for row in rows:
            if row and len(row) > 1:
                result.append({
                    "raw_date": str(row[1]),
                    "type": str(type(row[1]).__name__),
                    "status": str(row[12]) if len(row) > 12 else ""
                })
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/api/daily")
@login_required
def daily():
    return jsonify(data_store.get_daily_stats())

@app.route("/api/stats")
@login_required
def stats():
    return jsonify(data_store.get_stats())

@app.route("/api/export")
@login_required
def export():
    if not data_store.OUTPUT_FILE.exists():
        return "No data yet", 404
    import datetime
    return send_file(str(data_store.OUTPUT_FILE), as_attachment=True,
                     download_name=f"LAT_Report_{datetime.date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    print("\n" + "="*55)
    print("  Toronto Elections — LAT Data Collection System")
    print("="*55)
    data_store.init()
    # ── HTTPS / SSL ────────────────────────────────────────────────────────
    cert_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cert.pem")
    key_file  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "key.pem")

    if os.path.exists(cert_file) and os.path.exists(key_file):
        ssl_ctx  = (cert_file, key_file)
        protocol = "https"
        print("  HTTPS:     Enabled (cert.pem + key.pem found)")
    else:
        ssl_ctx  = None
        protocol = "http"
        print("  HTTPS:     Disabled - run GENERATE_CERT.bat first")

    hostname = os.environ.get("COMPUTERNAME", "localhost")
    print("\n  Local:     " + protocol + "://localhost:5000")
    print("  Network:   " + protocol + "://" + hostname + ":5000")
    print("  Dashboard: " + protocol + "://" + hostname + ":5000/dashboard")
    print("  Password:  " + APP_PASSWORD + "  (change via LAT_APP_PASSWORD env var)")
    print("\n  Press Ctrl+C to stop.\n")

    threading.Timer(1.5, lambda: webbrowser.open(protocol + "://localhost:5000")).start()
    app.run(host="0.0.0.0", port=5000, debug=False,
            use_reloader=False, threaded=True,
            ssl_context=ssl_ctx)
